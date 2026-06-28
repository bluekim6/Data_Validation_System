"""ExcelReader 단위 테스트 (1단계 완료 기준).

임시 .xlsx를 생성하여 시트 목록/헤더 인식/미리보기/예외 처리를 검증한다.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest

from src.dataio import (
    ExcelReader,
    InvalidFileFormatError,
    NoColumnsError,
    NoDataError,
)


@pytest.fixture()
def sample_xlsx(tmp_path: Path) -> Path:
    path = tmp_path / "maintenance.xlsx"
    df = pd.DataFrame(
        {
            "Data ID": ["EQ-001", "EQ-002"],
            "Maintenance Cycle": [12, "ABC"],
            "Criticality": ["High", "Very High"],
        }
    )
    df.to_excel(path, index=False, sheet_name="Plan")
    return path


def test_sheet_names(sample_xlsx: Path) -> None:
    reader = ExcelReader(sample_xlsx)
    assert reader.sheet_names() == ["Plan"]


def test_read_sheet_headers_and_values(sample_xlsx: Path) -> None:
    reader = ExcelReader(sample_xlsx)
    df = reader.read_sheet("Plan")
    assert list(df.columns) == ["Data ID", "Maintenance Cycle", "Criticality"]
    assert len(df) == 2
    assert df.iloc[0]["Data ID"] == "EQ-001"


def test_read_sheet_defaults_to_first(sample_xlsx: Path) -> None:
    reader = ExcelReader(sample_xlsx)
    df = reader.read_sheet()  # 시트명 생략 → 첫 시트
    assert len(df) == 2


def test_preview_limits_rows(sample_xlsx: Path) -> None:
    reader = ExcelReader(sample_xlsx)
    assert len(reader.preview(n=1)) == 1


def test_unsupported_format(tmp_path: Path) -> None:
    bad = tmp_path / "data.csv"
    bad.write_text("a,b\n1,2\n", encoding="utf-8")
    with pytest.raises(InvalidFileFormatError):
        ExcelReader(bad)


def test_missing_file(tmp_path: Path) -> None:
    with pytest.raises(InvalidFileFormatError):
        ExcelReader(tmp_path / "nope.xlsx")


def test_no_data_rows(tmp_path: Path) -> None:
    path = tmp_path / "empty.xlsx"
    pd.DataFrame(columns=["Data ID", "Criticality"]).to_excel(path, index=False)
    reader = ExcelReader(path)
    with pytest.raises(NoDataError):
        reader.read_sheet()


def test_no_header(tmp_path: Path) -> None:
    # 헤더 없이 데이터만 있는 시트 → 첫 행이 헤더로 잡히지만 모두 Unnamed가 되도록 구성
    path = tmp_path / "noheader.xlsx"
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = None
    ws["B1"] = None
    wb.save(path)

    reader = ExcelReader(path)
    with pytest.raises((NoColumnsError, NoDataError)):
        reader.read_sheet()
