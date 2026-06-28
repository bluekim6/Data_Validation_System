"""담당자별 리포트 생성 테스트 (PRD §5.7)."""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from src.models import ErrorRecord, ErrorType
from src.reporting import generate_reports, group_by_owner, write_report


def _errors() -> list[ErrorRecord]:
    return [
        ErrorRecord(data_id="EQ-001", row_number=3, column_name="Maintenance Cycle",
                    input_value="ABC", error_type=ErrorType.INVALID_DATA_TYPE,
                    error_message="숫자 형식으로 입력해 주세요.",
                    responsible_person="홍길동",
                    responsible_email="gildong.hong@company.com"),
        ErrorRecord(data_id="EQ-003", row_number=5, column_name="Criticality",
                    input_value="Very High", error_type=ErrorType.INVALID_LIST_VALUE,
                    error_message="High, Medium, Low 중 하나여야 합니다.",
                    responsible_person="홍길동",
                    responsible_email="gildong.hong@company.com"),
        ErrorRecord(data_id="EQ-002", row_number=4, column_name="Responsible Email",
                    input_value="bad", error_type=ErrorType.INVALID_EMAIL_FORMAT,
                    error_message="이메일 형식이 올바르지 않습니다.",
                    responsible_person="김철수",
                    responsible_email="chulsoo.kim@company.com"),
        # 이메일 없는 오류(매핑 오류 등)는 그룹화에서 제외되어야 함
        ErrorRecord(row_number=0, column_name="Responsible Email",
                    error_type=ErrorType.COLUMN_MAPPING_ERROR,
                    error_message="매핑 필요"),
    ]


def test_group_by_owner_excludes_no_email() -> None:
    groups = group_by_owner(_errors())
    emails = {g.email for g in groups}
    assert emails == {"gildong.hong@company.com", "chulsoo.kim@company.com"}
    gildong = next(g for g in groups if g.email == "gildong.hong@company.com")
    assert gildong.error_count == 2


def test_file_name_format() -> None:
    groups = group_by_owner(_errors())
    gildong = next(g for g in groups if g.person == "홍길동")
    name = gildong.file_name(date(2026, 6, 27))
    assert name == "Validation_Report_홍길동_2026-06-27.xlsx"


def test_write_report_creates_file_with_content(tmp_path: Path) -> None:
    groups = group_by_owner(_errors())
    gildong = next(g for g in groups if g.person == "홍길동")
    path = write_report(gildong, tmp_path, date(2026, 6, 27))

    assert path.exists()
    wb = load_workbook(path)
    ws = wb.active
    flat = [c.value for row in ws.iter_rows() for c in row]
    assert "홍길동" in flat
    assert "gildong.hong@company.com" in flat
    assert "EQ-001" in flat
    assert "EQ-003" in flat


def test_generate_reports_one_file_per_owner(tmp_path: Path) -> None:
    out = generate_reports(_errors(), tmp_path)
    assert len(out) == 2
    for _, path in out:
        assert path.exists()
