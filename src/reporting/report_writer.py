"""담당자별 오류 리포트 생성 (PRD §5.7).

ValidationResult의 ErrorRecord를 Responsible Email 기준으로 그룹화하여 담당자별
Excel 리포트를 만든다. 원본 파일은 변경하지 않고 별도 파일로 생성한다(PRD §14.4).

리포트 구성(PRD §5.7):
  1) 리포트 요약  2) 담당자 정보  3) 오류 건수 요약
  4) 오류 상세 목록  5) 수정 요청 코멘트  6) 재제출 안내
"""

from __future__ import annotations

import re
from collections import Counter
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field

from ..models import ErrorRecord

UNASSIGNED = "(미지정)"
_HEADER_FILL = PatternFill("solid", fgColor="DDDDDD")
_TITLE_FONT = Font(bold=True, size=14)
_BOLD = Font(bold=True)
_DETAIL_HEADERS = ["Data ID", "Row No.", "Column", "Input Value", "Error Type", "Comment"]

_RESUBMIT_NOTICE = (
    "위 오류 항목을 수정하신 후, 동일한 파일을 재제출해 주시기 바랍니다. "
    "문의 사항은 데이터 검토 담당자에게 연락 주세요."
)


class OwnerReport(BaseModel):
    """담당자 한 명에 대한 오류 묶음."""

    person: str
    email: str
    errors: list[ErrorRecord] = Field(default_factory=list)

    @property
    def error_count(self) -> int:
        return len(self.errors)

    def file_name(self, report_date: date | None = None) -> str:
        """리포트 파일명 자동 생성 (PRD §5.7)."""
        report_date = report_date or date.today()
        safe = _sanitize(self.person)
        return f"Validation_Report_{safe}_{report_date.isoformat()}.xlsx"


def group_by_owner(errors: list[ErrorRecord]) -> list[OwnerReport]:
    """오류를 담당자 이메일 기준으로 그룹화한다.

    이메일이 없는 오류(예: 컬럼 매핑 오류)는 발송 대상이 아니므로 제외한다.
    """
    groups: dict[str, OwnerReport] = {}
    for err in errors:
        if not err.responsible_email:
            continue
        key = err.responsible_email
        if key not in groups:
            groups[key] = OwnerReport(
                person=err.responsible_person or UNASSIGNED, email=key
            )
        groups[key].errors.append(err)
    return list(groups.values())


def write_report(
    report: OwnerReport,
    out_dir: str | Path,
    report_date: date | None = None,
) -> Path:
    """담당자 리포트를 Excel로 저장하고 경로를 반환한다."""
    report_date = report_date or date.today()
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / report.file_name(report_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Validation Report"

    # 1) 리포트 요약 제목
    ws["A1"] = "유지보수 데이터 검증 오류 리포트"
    ws["A1"].font = _TITLE_FONT

    # 2) 담당자 정보
    info = [
        ("담당자", report.person),
        ("이메일", report.email),
        ("오류 건수", report.error_count),
        ("생성일", report_date.isoformat()),
    ]
    row = 3
    for label, value in info:
        ws.cell(row=row, column=1, value=label).font = _BOLD
        ws.cell(row=row, column=2, value=value)
        row += 1

    # 3) 오류 유형별 건수
    row += 1
    ws.cell(row=row, column=1, value="오류 유형별 건수").font = _BOLD
    row += 1
    for etype, count in Counter(e.error_type.value for e in report.errors).items():
        ws.cell(row=row, column=1, value=etype)
        ws.cell(row=row, column=2, value=count)
        row += 1

    # 4) 오류 상세 목록
    row += 1
    ws.cell(row=row, column=1, value="오류 상세").font = _BOLD
    row += 1
    for col, header in enumerate(_DETAIL_HEADERS, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = _BOLD
        cell.fill = _HEADER_FILL
    row += 1
    for err in report.errors:
        values = [
            err.data_id,
            err.row_number,
            err.column_name,
            err.input_value,
            err.error_type.value,
            err.error_message,
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    # 5) 수정 요청 코멘트 / 6) 재제출 안내
    row += 1
    note = ws.cell(row=row, column=1, value=_RESUBMIT_NOTICE)
    note.alignment = Alignment(wrap_text=True)

    _autofit(ws)
    wb.save(path)
    return path


def generate_reports(
    errors: list[ErrorRecord],
    out_dir: str | Path,
    report_date: date | None = None,
) -> list[tuple[OwnerReport, Path]]:
    """모든 담당자 리포트를 생성하고 (리포트, 경로) 목록을 반환한다."""
    results: list[tuple[OwnerReport, Path]] = []
    for report in group_by_owner(errors):
        path = write_report(report, out_dir, report_date)
        results.append((report, path))
    return results


def _sanitize(name: str) -> str:
    """파일명에 안전하도록 공백/특수문자를 치환한다."""
    return re.sub(r"[^\w가-힣.-]+", "_", name).strip("_") or "unknown"


def _autofit(ws) -> None:
    for column_cells in ws.columns:
        length = max(
            (len(str(c.value)) for c in column_cells if c.value is not None),
            default=0,
        )
        letter = column_cells[0].column_letter
        ws.column_dimensions[letter].width = min(max(length + 2, 10), 60)
